
import requests
from openpyxl import load_workbook

'''
Cyber Vision Info
'''
center_token = "ics-9591546cceca88d78bd8b6d87241bd819128c14d-0b08250afd8dfca4b40094019540e65c0b2fcc81"
center_ip = "10.70.138.125"
'''
read excel
'''
wb = load_workbook("Factory_Inventory.xlsx")
#print (type(wb))
#print (wb.sheetnames)
sheetname = wb.sheetnames[0]
ws = wb[sheetname]

for row in range(2,5):
    '''
    read excel line and insert list in list
    '''
    group = ws.cell(row, 1).value
    IP_Address = ws.cell(row, 2).value
    Type = ws.cell(row, 3).value
    Host_Name = ws.cell(row, 4).value
    Location = ws.cell(row, 5).value
    Owner = ws.cell(row, 6).value
    Phone = ws.cell(row, 7).value
    #print ("###_L_26", group, IP_Address, Type, Host_Name, Location, Owner, Phone)
    insert_data = [ ["Type", Type], ["Host_Name", Host_Name], ["Location", Location], ["Owner", Owner], ["Phone", Phone] ]
    print ("\n\n\n#######_L_30", insert_data)
    #
    # getting ID of object based group and IP_Address
    #
    center_port = "443"
    center_base_url = "api/3.0"
    headers = {"x-token-id": center_token}
    r_get = requests.get(f"https://{center_ip}:{center_port}/{center_base_url}/components", headers=headers,
                         verify=False)
    raw_json_data = r_get.json()

    for component in raw_json_data:
        '''
        insert into cybervision line by line using API
        '''
        if component["group"] != None:
            # find the component ID
            print("###_L_42", component["group"]["label"], group, component["ip"], IP_Address)
            if component["group"]["label"] == group and component["ip"] == IP_Address:
                print("#############")
                print("bingo", group, IP_Address, component["id"])
                target_ID = component["id"]
                #insert
                for i in range (0,5):
                    json_data = {'label': insert_data[i][0], 'value': insert_data[i][1] }
                    print ("####_L_54_to_insert", json_data)
                    r_get = requests.post(f"https://{center_ip}:{center_port}/{center_base_url}/components/{target_ID}/usersProperties",headers=headers, json=json_data, verify=False)
                    print("####_L_55", r_get,"\n\n\n")
